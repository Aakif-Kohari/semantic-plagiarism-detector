import inspect
import io
import os
from datetime import datetime, timezone

import openpyxl
import pandas as pd

from src.utils.bulk_export import export_incidents_xlsx_stream
from src.utils.excel_export import (
    build_similarity_workbook,
    export_similarity_matrix_to_excel,
    export_similarity_matrix_to_temp_file,
    generate_csv_matrix_stream,
    generate_tsv_matrix_stream,
)


def test_generate_csv_matrix_stream():
    # Setup test DataFrame
    data = {
        "DocA.txt": [1.0, 0.85, 0.12],
        "DocB.txt": [0.85, 1.0, 0.45],
        "DocC.txt": [0.12, 0.45, 1.0],
    }
    df = pd.DataFrame(data, index=["DocA.txt", "DocB.txt", "DocC.txt"])

    # Test 1: Return type is a Generator
    stream = generate_csv_matrix_stream(df)
    assert inspect.isgenerator(stream)

    # Test 2: Verify chunk output
    chunks = list(stream)
    assert len(chunks) == len(df) + 1  # 1 header row + 3 data rows

    # Verify header line
    assert chunks[0].strip() == "Document,DocA.txt,DocB.txt,DocC.txt"

    # Verify data lines
    assert chunks[1].strip() == "DocA.txt,1.0,0.85,0.12"
    assert chunks[2].strip() == "DocB.txt,0.85,1.0,0.45"
    assert chunks[3].strip() == "DocC.txt,0.12,0.45,1.0"

    # Test 3: Verify complete CSV reconstruction matches Expected CSV output
    full_csv = "".join(chunks)
    reconstructed_df = pd.read_csv(io.StringIO(full_csv), index_col=0)
    pd.testing.assert_frame_equal(df, reconstructed_df, check_names=False)


def test_generate_tsv_matrix_stream():
    # Setup test DataFrame
    data = {
        "DocA.txt": [1.0, 0.85, 0.12],
        "DocB.txt": [0.85, 1.0, 0.45],
        "DocC.txt": [0.12, 0.45, 1.0],
    }
    df = pd.DataFrame(data, index=["DocA.txt", "DocB.txt", "DocC.txt"])

    # Test 1: Return type is a Generator
    stream = generate_tsv_matrix_stream(df)
    assert inspect.isgenerator(stream)

    # Test 2: Verify chunk output
    chunks = list(stream)
    assert len(chunks) == len(df) + 1  # 1 header row + 3 data rows

    # Verify header line with tabs
    assert chunks[0].strip() == "Document\tDocA.txt\tDocB.txt\tDocC.txt"

    # Verify data lines with tabs
    assert chunks[1].strip() == "DocA.txt\t1.0\t0.85\t0.12"
    assert chunks[2].strip() == "DocB.txt\t0.85\t1.0\t0.45"
    assert chunks[3].strip() == "DocC.txt\t0.12\t0.45\t1.0"

    # Test 3: Verify complete TSV reconstruction matches Expected TSV output
    full_tsv = "".join(chunks)
    reconstructed_df = pd.read_csv(io.StringIO(full_tsv), sep="\t", index_col=0)
    pd.testing.assert_frame_equal(df, reconstructed_df, check_names=False)



def test_build_similarity_workbook_metadata_properties():
    """Verify build_similarity_workbook populates document title, creator, and created timestamp (#3438)."""
    df = pd.DataFrame({"Doc1.txt": [1.0]}, index=["Doc1.txt"])
    before = datetime.now(timezone.utc)
    wb = build_similarity_workbook(df)
    after = datetime.now(timezone.utc)

    assert wb.properties.title == "Semantic Plagiarism Similarity Report"
    assert wb.properties.creator == "Semantic Plagiarism Detector"
    assert wb.properties.created is not None
    assert isinstance(wb.properties.created, datetime)
    assert before <= wb.properties.created <= after


def test_export_similarity_matrix_to_excel_persists_metadata():
    """Verify export_similarity_matrix_to_excel persists metadata in the saved XLSX file (#3438)."""
    df = pd.DataFrame(
        {"DocA.txt": [1.0, 0.8], "DocB.txt": [0.8, 1.0]},
        index=["DocA.txt", "DocB.txt"],
    )
    xlsx_bytes = export_similarity_matrix_to_excel(df)
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 0

    # Load back with openpyxl to inspect file properties
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.properties.title == "Semantic Plagiarism Similarity Report"
    assert wb.properties.creator == "Semantic Plagiarism Detector"
    assert wb.properties.created is not None


def test_export_incidents_xlsx_stream_persists_metadata():
    """Verify export_incidents_xlsx_stream sets title, creator, and created metadata (#3438)."""
    incidents = [
        {
            "incident_id": "INC-001",
            "document_a": "Essay1.docx",
            "document_b": "Essay2.docx",
            "similarity_score": 0.88,
            "severity_rank": "High",
            "review_status": "Pending",
            "date_flagged": "2026-08-25",
        }
    ]
    xlsx_bytes = export_incidents_xlsx_stream(incidents)
    assert isinstance(xlsx_bytes, bytes)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.properties.title == "Semantic Plagiarism Similarity Report"
    assert wb.properties.creator == "Semantic Plagiarism Detector"
    assert wb.properties.created is not None


def test_build_similarity_workbook_write_only_flag():
    """Verify write_only flag controls openpyxl.Workbook write_only mode (#3435)."""
    df = pd.DataFrame({"DocA.txt": [1.0]}, index=["DocA.txt"])

    wb_standard = build_similarity_workbook(df, write_only=False)
    assert wb_standard.write_only is False

    wb_stream = build_similarity_workbook(df, write_only=True)
    assert wb_stream.write_only is True


def test_write_only_export_similarity_matrix_to_excel_roundtrip():
    """Verify write_only=True produces valid XLSX with identical data and metadata (#3435)."""
    data = {
        "DocA.txt": [1.0, 0.85, 0.12],
        "DocB.txt": [0.85, 1.0, 0.45],
        "DocC.txt": [0.12, 0.45, 1.0],
    }
    df = pd.DataFrame(data, index=["DocA.txt", "DocB.txt", "DocC.txt"])

    xlsx_bytes = export_similarity_matrix_to_excel(df, threshold=0.60, write_only=True)
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 0

    # Load the generated XLSX back with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert "Similarity Matrix" in wb.sheetnames
    ws = wb["Similarity Matrix"]

    # Verify header
    headers = [cell.value for cell in ws[1]]
    assert headers == ["Document", "DocA.txt", "DocB.txt", "DocC.txt"]

    # Verify rows
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 4  # Header + 3 data rows
    assert rows[1][0] == "DocA.txt"
    assert rows[1][1] == 1.0
    assert rows[1][2] == 0.85
    assert rows[1][3] == 0.12

    # Verify document properties
    assert wb.properties.title == "Semantic Plagiarism Similarity Report"
    assert wb.properties.creator == "Semantic Plagiarism Detector"
    assert wb.properties.created is not None


def test_write_only_export_similarity_matrix_to_temp_file():
    """Verify export_similarity_matrix_to_temp_file works with write_only=True (#3435)."""
    df = pd.DataFrame(
        {"Doc1.txt": [1.0, 0.5], "Doc2.txt": [0.5, 1.0]},
        index=["Doc1.txt", "Doc2.txt"],
    )
    temp_file = export_similarity_matrix_to_temp_file(df, write_only=True)
    try:
        assert os.path.exists(temp_file)
        assert temp_file.endswith(".xlsx")

        wb = openpyxl.load_workbook(temp_file)
        ws = wb.active
        assert ws.title == "Similarity Matrix"
        assert ws.cell(row=1, column=1).value == "Document"
    finally:
        if os.path.exists(temp_file):
            os.remove(temp_file)


def test_write_only_large_matrix_export():
    """Verify write_only streams larger matrices without memory or structural errors (#3435)."""
    dim = 25
    doc_names = [f"Student_Document_{i:03d}.docx" for i in range(dim)]
    matrix_data = {doc: [0.75 for _ in range(dim)] for doc in doc_names}
    df = pd.DataFrame(matrix_data, index=doc_names)

    xlsx_bytes = export_similarity_matrix_to_excel(df, threshold=0.70, write_only=True)
    assert len(xlsx_bytes) > 0

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Similarity Matrix"]
    assert ws.max_row == dim + 1
    assert ws.max_column == dim + 1

